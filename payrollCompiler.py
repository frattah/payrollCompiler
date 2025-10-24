import numpy as np
import logging
import pytesseract
from pdf2image import convert_from_path
import sys
import cv2
import re
from PIL import Image
from openpyxl import load_workbook
from pathlib import Path
import concurrent.futures
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime

######################################################################################################################
#                                                                                                                    #
#                                                 PROFILES                                                           #
#                                                                                                                    #
######################################################################################################################

MACCHINISTA = ["attendances", "vacancies","0131","0169","0170","0412","holidays","0421","0457","0547","0790","0791","0792","0964","0965","0966","0987","0988","0991","0992","0AD0","0AD1"]
MANUTENZIONE = ["attendances", "vacancies","tickets","0131","holidays","0412","0457","0470","0482","0496","0584","0686","0687","0423"]

######################################################################################################################
#                                                                                                                    #
#                                               PARAMETERS                                                           #
#                                                                                                                    #
######################################################################################################################

WRITING_PARAMETERS = MANUTENZIONE
CODES = ["0131","0169","0170","0200","0202","0203","0205","0206","0207","0210","0293","0299","0352","0353",
        "0366","0412","0421","0423","0457","0470","0482","0496","0547","0584","0686","0687","0790","0791",
        "0792","0964","0965","0966","0987","0988","0991","0992","0AD0","0AD1"]
HOLIDAYS = ["0200","0202","0203","0205","0206","0207","0210","0352","0353","0366"]
TICKETS = ["0293","0299"]
MONTHS = {
    "Gennaio": 1,
    "Febbraio": 2,
    "Marzo": 3,
    "Aprile": 4,
    "Maggio": 5,
    "Giugno": 6,
    "Luglio": 7,
    "Agosto": 8,
    "Settembre": 9,
    "Ottobre": 10,
    "Novembre": 11,
    "Dicembre": 12
}
Y_STARTING_CELL = 5
X_STARTING_CELL = 8
FIRST_YEAR = 2007
excel_path = sys.argv[1]
MAX_WORKERS = 8 

######################################################################################################################
#                                                                                                                    #
#                                                     CODE                                                           #
#                                                                                                                    #
######################################################################################################################

class Payroll:
    def __init__(self, year, month):
        self.year = year
        self.month = month
        self.pay_elements = {}

    def set_pay_element(self, name, value):
        if name in self.pay_elements:
            self.pay_elements[name] += value
        else:
            self.pay_elements[name] = value

    def get_pay_element(self, name):
        if name not in self.pay_elements:
            return 0
        return self.pay_elements[name]

    def compute_derivates(self):
        self.pay_elements["holidays"] = sum(self.pay_elements[key] for key in self.pay_elements if key in HOLIDAYS)

        # Choose the right ticket value
        if self.year > 2023 or (self.year == 2023 and MONTHS[self.month] >= 10):
            ticket_value = 10.50
        elif self.year < 2012 or (self.year == 2012 and MONTHS[self.month] < 10):
            ticket_value = 0
        else:
            ticket_value = 7.30

        # Multiply ticket value to ticket parameter
        if "0293" in self.pay_elements:
            self.pay_elements["tickets"] = self.pay_elements["0293"] * ticket_value
        elif "0299" in self.pay_elements:
            self.pay_elements["tickets"] = self.pay_elements["0299"] * ticket_value

    def write_on_spreadsheet(self, sheet):
        years_elapsed = self.year - FIRST_YEAR
        cumulative_offset = (years_elapsed // 2) * 2 # Use it for printable excel version

        for i, pay_element in enumerate(WRITING_PARAMETERS):
            row_index = (
                Y_STARTING_CELL + 
                (self.year - FIRST_YEAR) * (len(WRITING_PARAMETERS) + 8) + 
                i
            )
 
            sheet.cell(
                row=row_index, 
                column=X_STARTING_CELL + MONTHS[self.month], 
                value=self.pay_elements[pay_element] if pay_element in self.pay_elements else 0.00
            )

# ------------------------------------------------------------------------------------------------------------------

pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

def process_pdf(pdf_file):

    pages = convert_from_path(pdf_file, dpi=300, poppler_path=r"C:\Users\framo\Downloads\Release-25.07.0-0\poppler-25.07.0\Library\bin")

    # Read pdf content after converting it to image
    text = ""
    for page_num, page in enumerate(pages, start=1):
        img = page.convert("RGB")
        custom_config = r'--psm 4 -l ita'
        page_text = pytesseract.image_to_string(img, config=custom_config)
        text += page_text

    current_year = datetime.now().year
    year_pattern = rf"\s+(200[7-9]|20[0-{str(current_year)[2]}][0-9])"
    pattern = r"(" + "|".join(MONTHS.keys()) + r")[^\n]{0,20}?(" + year_pattern + r")"
    
    # Find month and year of the payroll
    match = re.search(
        pattern,
        text,
        re.IGNORECASE | re.DOTALL
    ) # Case insensitive and capture also '\n'

    if match:
        month = match.group(1)
        year = int(match.group(2))

    payroll = Payroll(year, month)

    # Find attendance and vacancies
    pattern = r"Ferie anno.*"
    match = re.search(pattern, text, re.DOTALL | re.IGNORECASE)

    if match:
        numbers = re.findall(r"\d{1,2},\d{2}", match.group(0))
        attendances = float(numbers[0].replace(',', '.')) if len(numbers) >= 1 else 0.0
        vacancies = float(numbers[2].replace(',', '.')) if len(numbers) >= 3 else 0.0
    else:
        attendances = 0.0
        vacancies = 0.0
    payroll.set_pay_element("attendances", attendances)
    payroll.set_pay_element("vacancies", vacancies)

    # Find all pay elements in the payroll
    for code in CODES:
        if code in TICKETS:
            # Pattern: codice all'inizio della riga, poi qualsiasi cosa, poi cattura il primo numero con eventuale virgola
            pattern = rf"^{code}\b.*?\s(\d+(?:,\d+)?)"
        else:
            # Per gli altri codici, continua a prendere l'ultimo numero della riga
            pattern = rf"^{code}\b.*?(\d+(?:[.,]\d+)?)$"

        for match in re.finditer(pattern, text, re.MULTILINE):
            value = float(match.group(1).replace(",", "."))
            payroll.set_pay_element(code, value)

    print(f"Processed {payroll.month} {payroll.year}")
    return payroll

def process_payrolls(sheet):
    all_payrolls = []
    pdf_files = list(Path(".").rglob("*.pdf"))

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {executor.submit(process_pdf, str(pdf_file)): pdf_file for pdf_file in pdf_files}

        for future in concurrent.futures.as_completed(futures):
            pdf_file = futures[future]
            try:
                payroll = future.result()
                all_payrolls.append(payroll)
            except Exception as e:
                print(f"Errore con {pdf_file}: {e}")

    for payroll in all_payrolls:
        payroll.compute_derivates()
        payroll.write_on_spreadsheet(sheet)


wb = load_workbook(excel_path)
sheet = wb[wb.sheetnames[0]]

process_payrolls(sheet)

wb.save("buste_modificato.xlsx")



